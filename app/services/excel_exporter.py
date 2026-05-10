# ============================================================
#  excel_exporter.py — Convert Extracted Data to Excel File
#
#  Takes the dictionary of extracted fields and creates
#  a nicely formatted .xlsx file using pandas + openpyxl.
#
#  Output Excel has:
#    - Sheet 1: "Extracted Data"  → main fields, one row per doc
#    - Sheet 2: "Line Items"      → receipt line items (if any)
# ============================================================

import os
import uuid
from datetime import datetime
import pandas as pd


def export_to_excel(records, output_folder):
    """
    Convert a list of extracted field dicts into a formatted Excel file.

    Example input:
      records = [
        {"doc_type": "invoice", "date": "2024-03-12", "amount": 5000.0, "vendor": "ABC Ltd"},
        {"doc_type": "receipt", "date": "2024-03-13", "amount": 320.0}
      ]

    Returns: Full path to the created .xlsx file
    """
    if not records:
        raise ValueError("No records to export!")

    # Generate a unique filename using timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = uuid.uuid4().hex[:6]
    filename  = f"extracted_{timestamp}_{unique_id}.xlsx"
    filepath  = os.path.join(output_folder, filename)

    os.makedirs(output_folder, exist_ok=True)

    # Separate flat fields from nested line items
    flat_records  = []
    all_line_items = []

    for i, record in enumerate(records, 1):
        # Flat fields (simple key → value)
        flat = {}
        flat["S.No"] = i
        for key, value in record.items():
            # Skip private keys (start with _) and nested lists
            if not key.startswith("_") and not isinstance(value, (list, dict)):
                flat[key] = value
        flat_records.append(flat)

        # Handle line items separately (receipts have item lists)
        for item in (record.get("line_items") or []):
            item_row = dict(item)
            item_row["Document #"] = i
            all_line_items.append(item_row)

    # Create main DataFrame
    df_main = _organize_columns(pd.DataFrame(flat_records))

    # Write to Excel
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Sheet 1: Main extracted data
        df_main.to_excel(writer, sheet_name="Extracted Data", index=False)

        # Sheet 2: Line items (only if we have any)
        if all_line_items:
            df_items = pd.DataFrame(all_line_items)
            df_items.to_excel(writer, sheet_name="Line Items", index=False)

        # Apply nice formatting to both sheets
        _apply_formatting(writer, "Extracted Data", df_main)
        if all_line_items:
            _apply_formatting(writer, "Line Items", pd.DataFrame(all_line_items))

    print(f"Excel file created: {filepath}")
    return filepath


def _organize_columns(df):
    """Put important columns first, in a logical order."""
    priority_columns = [
        "S.No", "doc_type", "date", "due_date",
        "invoice_no", "receipt_no", "account_no",
        "vendor", "person", "amount", "tax_amount",
        "closing_balance", "opening_balance",
        "payment_mode", "gstin", "location", "source_file"
    ]
    # Only include columns that actually exist in our data
    ordered = [col for col in priority_columns if col in df.columns]
    # Add any remaining columns at the end
    remaining = [col for col in df.columns if col not in ordered]
    return df[ordered + remaining]


def _apply_formatting(writer, sheet_name, df):
    """
    Make the Excel sheet look professional:
    - Dark blue header row with white text
    - Auto-fit column widths
    - Alternating row colors
    - Freeze the header row
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        ws = writer.sheets[sheet_name]

        # Style the header row (row 1)
        header_fill = PatternFill("solid", fgColor="1F4E79")  # Dark blue
        header_font = Font(bold=True, color="FFFFFF", size=11) # White text

        for col_idx, column_name in enumerate(df.columns, 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill      = header_fill
            header_cell.font      = header_font
            header_cell.alignment = Alignment(horizontal="center")

            # Auto-fit column width based on content
            max_length = max(
                len(str(column_name)),
                *[len(str(ws.cell(row=r, column=col_idx).value or ""))
                  for r in range(2, min(ws.max_row + 1, 100))]
            )
            ws.column_dimensions[header_cell.column_letter].width = min(max_length + 4, 40)

        # Alternating row colors (light blue for even rows)
        light_blue = PatternFill("solid", fgColor="EBF3FB")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = light_blue

        # Set header row height and freeze it
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"  # Header stays visible when scrolling

    except Exception as e:
        print(f"Formatting warning: {e}")  # Don't crash if formatting fails
